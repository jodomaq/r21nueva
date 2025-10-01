import hashlib
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Dict, Iterable, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from fpdf import FPDF
from openpyxl import Workbook
from sqlalchemy import String, cast, distinct, func
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.orm import selectinload
from sqlmodel import Session, select

from .. import models, schemas
from ..database import get_session

router = APIRouter(prefix="/dashboard", tags=["dashboard"])

ROLE_LABELS = {
    1: "Coordinación Estatal",
    2: "Delegación Regional",
    3: "Coordinación Distrital",
    4: "Coordinación Municipal",
    5: "Coordinación Seccional",
    6: "Presidencia de Comité",
}

MUNICIPALITY_TARGET = 113


def _role_label(role: Optional[int]) -> str:
    if role is None:
        return "Rol sin definir"
    return ROLE_LABELS.get(role, "Rol sin definir")


def _safe_scalar(session: Session, statement, default: int | float = 0) -> int | float:
    try:
        result = session.exec(statement)
        if hasattr(result, "scalar_one_or_none"):
            value = result.scalar_one_or_none()
        else:
            value = result.one_or_none()
        if value is None:
            return default
        if hasattr(value, "_mapping"):
            mapping = value._mapping
            value = next(iter(mapping.values()), None)
        elif isinstance(value, (list, tuple)):
            value = value[0] if value else None
        return value if value is not None else default
    except SQLAlchemyError:
        return default


def _normalize_upload_path(filename: str) -> str:
    normalized = filename.replace("\\", "/") if filename else ""
    if normalized.startswith("/uploads/"):
        return normalized
    return f"/uploads/{normalized}" if normalized else normalized


def _latin1(text: Optional[str]) -> str:
    if text is None:
        return ""
    return text.encode("latin-1", "ignore").decode("latin-1")


def _serialize_committees(session: Session, committees: Iterable[models.Committee]) -> List[schemas.CommitteeDashboardOut]:
    committees_list = list(committees)
    if not committees_list:
        return []

    owner_emails = {c.owner_id for c in committees_list if c.owner_id}
    owner_map: Dict[str, models.User] = {}
    if owner_emails:
        owners = session.exec(
            select(models.User).where(models.User.email.in_(owner_emails))
        ).all()
        owner_map = {owner.email: owner for owner in owners}

    section_ids: set[int] = set()
    for committee in committees_list:
        if committee.section_number:
            try:
                section_ids.add(int(committee.section_number))
            except ValueError:
                continue
    section_map: Dict[int, models.Seccion] = {}
    if section_ids:
        try:
            sections = session.exec(
                select(models.Seccion).where(models.Seccion.id.in_(section_ids))
            ).all()
            section_map = {section.id: section for section in sections if section.id is not None}
        except SQLAlchemyError:
            section_map = {}

    payload: List[schemas.CommitteeDashboardOut] = []
    for committee in committees_list:
        owner = owner_map.get(committee.owner_id)
        section_ref = None
        if committee.section_number:
            try:
                section_model = section_map.get(int(committee.section_number))
            except ValueError:
                section_model = None
            section_ref = (
                schemas.SectionRef.model_validate(section_model)
                if section_model is not None
                else None
            )
        admin_ref = (
            schemas.AdministrativeUnitRef.model_validate(committee.administrative_unit)
            if committee.administrative_unit is not None
            else None
        )
        members = [
            schemas.CommitteeMemberOut.model_validate(member)
            for member in getattr(committee, "members", [])
        ]
        documents = [
            schemas.DocumentWithUrl(
                id=document.id,
                filename=document.filename,
                original_name=document.original_name,
                content_type=document.content_type,
                size=document.size,
                created_at=document.created_at,
                url=_normalize_upload_path(document.filename),
            )
            for document in getattr(committee, "documents", [])
        ]
        payload.append(
            schemas.CommitteeDashboardOut(
                id=committee.id,
                name=committee.name,
                section_number=committee.section_number,
                type=committee.type,
                owner_id=committee.owner_id,
                owner_name=owner.name if owner else None,
                created_at=committee.created_at,
                presidente=committee.presidente,
                email=committee.email,
                clave_afiliacion=committee.clave_afiliacion,
                telefono=committee.telefono,
                administrative_unit=admin_ref,
                section=section_ref,
                members=members,
                documents=documents,
                total_members=len(members),
            )
        )
    return payload


def _load_committees(session: Session, committee_id: Optional[int] = None) -> List[models.Committee]:
    statement = (
        select(models.Committee)
        .options(
            selectinload(models.Committee.members),
            selectinload(models.Committee.documents),
            selectinload(models.Committee.administrative_unit),
        )
        .order_by(models.Committee.created_at.desc())
    )
    if committee_id is not None:
        statement = statement.where(models.Committee.id == committee_id)
    return session.exec(statement).all()


@router.get("/attendance", response_model=List[schemas.AttendanceOut])
def list_attendance(session: Session = Depends(get_session)) -> List[schemas.AttendanceOut]:
    statement = select(models.Attendance).order_by(models.Attendance.created_at.desc())
    return session.exec(statement).all()


@router.get("/attendance/map", response_model=List[schemas.AttendanceMapPoint])
def list_attendance_map(session: Session = Depends(get_session)) -> List[schemas.AttendanceMapPoint]:
    statement = (
        select(models.Attendance)
        .where(
            models.Attendance.latitude.is_not(None),
            models.Attendance.longitude.is_not(None)
        )
        .order_by(models.Attendance.created_at.desc())
    )
    records = session.exec(statement).all()
    return [
        schemas.AttendanceMapPoint(
            id=record.id,
            name=record.name,
            email=record.email,
            latitude=record.latitude,
            longitude=record.longitude,
            created_at=record.created_at,
        )
        for record in records
    ]


@router.get("/committee-stats", response_model=schemas.CommitteeStatsResponse)
def committee_stats(session: Session = Depends(get_session)) -> schemas.CommitteeStatsResponse:
    user_stmt = (
        select(
            models.Committee.owner_id.label("owner_email"),
            func.count(models.Committee.id).label("total"),
            models.User.id.label("user_id"),
            models.User.name.label("user_name"),
        )
        .select_from(models.Committee)
        .outerjoin(models.User, models.User.email == models.Committee.owner_id)
        .group_by(models.Committee.owner_id, models.User.id, models.User.name)
        .order_by(func.count(models.Committee.id).desc())
    )
    by_user = [
        schemas.CommitteeOwnerStat(
            owner_email=row.owner_email or "sin-correo",
            owner_name=row.user_name,
            owner_id=row.user_id,
            total=row.total or 0,
        )
        for row in session.exec(user_stmt).all()
    ]

    try:
        section_stmt = (
            select(
                models.Committee.section_number.label("section_number"),
                func.count(models.Committee.id).label("total"),
                models.Seccion.nombre_municipio.label("municipio"),
            )
            .select_from(models.Committee)
            .outerjoin(
                models.Seccion,
                cast(models.Seccion.id, String) == models.Committee.section_number,
            )
            .where(
                models.Committee.section_number.is_not(None),
                models.Committee.section_number != "",
            )
            .group_by(models.Committee.section_number, models.Seccion.nombre_municipio)
            .order_by(func.count(models.Committee.id).desc())
        )
        section_rows = session.exec(section_stmt).all()
    except SQLAlchemyError:
        section_rows = []

    by_section = [
        schemas.CommitteeLocationStat(
            code=row.section_number,
            label=f"Sección {row.section_number}" if row.section_number else "Sin sección",
            municipality=row.municipio,
            total=row.total or 0,
        )
        for row in section_rows
    ]

    try:
        municipality_stmt = (
            select(
                models.Seccion.nombre_municipio.label("municipio"),
                func.count(models.Committee.id).label("total"),
            )
            .select_from(models.Committee)
            .outerjoin(
                models.Seccion,
                cast(models.Seccion.id, String) == models.Committee.section_number,
            )
            .where(
                models.Committee.section_number.is_not(None),
                models.Committee.section_number != "",
                models.Seccion.nombre_municipio.is_not(None),
            )
            .group_by(models.Seccion.nombre_municipio)
            .order_by(func.count(models.Committee.id).desc())
        )
        municipality_rows = session.exec(municipality_stmt).all()
    except SQLAlchemyError:
        municipality_rows = []

    by_municipality = [
        schemas.CommitteeLocationStat(
            code=row.municipio,
            label=row.municipio or "Sin municipio",
            municipality=row.municipio,
            total=row.total or 0,
        )
        for row in municipality_rows
    ]

    type_stmt = (
        select(
            models.Committee.type.label("type"),
            func.count(models.Committee.id).label("total"),
        )
        .group_by(models.Committee.type)
        .order_by(func.count(models.Committee.id).desc())
    )
    by_type = [
        schemas.CommitteeTypeStat(type=row.type or "Sin tipo", total=row.total or 0)
        for row in session.exec(type_stmt).all()
    ]

    return schemas.CommitteeStatsResponse(
        by_user=by_user,
        by_section=by_section,
        by_municipality=by_municipality,
        by_type=by_type,
    )


@router.get("/administrative-tree", response_model=List[schemas.AdministrativeUnitNode])
def administrative_tree(session: Session = Depends(get_session)) -> List[schemas.AdministrativeUnitNode]:
    units = session.exec(select(models.AdministrativeUnit)).all()
    if not units:
        return []

    assignments = session.exec(
        select(models.UserAssignment)
        .options(selectinload(models.UserAssignment.user))
    ).all()

    assignments_by_unit: Dict[int, List[schemas.AssignmentUserSummary]] = defaultdict(list)
    for assignment in assignments:
        if assignment.administrative_unit_id is None:
            continue
        user = assignment.user
        if user is None:
            user_email = f"usuario{assignment.user_id}@desconocido.local"
            user_name = f"Usuario {assignment.user_id}"
            user_id = assignment.user_id
        else:
            user_email = user.email
            user_name = user.name
            user_id = user.id
        assignments_by_unit[assignment.administrative_unit_id].append(
            schemas.AssignmentUserSummary(
                user_id=user_id,
                user_name=user_name,
                user_email=user_email,
                role=assignment.role,
                role_label=_role_label(assignment.role),
            )
        )

    children_map: Dict[Optional[int], List[models.AdministrativeUnit]] = defaultdict(list)
    for unit in units:
        children_map[unit.parent_id].append(unit)

    def build_node(unit: models.AdministrativeUnit) -> schemas.AdministrativeUnitNode:
        children = [build_node(child) for child in sorted(children_map.get(unit.id, []), key=lambda u: u.name.lower())]
        return schemas.AdministrativeUnitNode(
            id=unit.id,
            name=unit.name,
            code=unit.code,
            unit_type=unit.unit_type,
            assignments=assignments_by_unit.get(unit.id, []),
            children=children,
        )

    roots = sorted(children_map.get(None, []), key=lambda u: u.name.lower())
    return [build_node(root) for root in roots]


@router.get("/user-assignments", response_model=List[schemas.UserAssignmentRow])
def list_user_assignments(session: Session = Depends(get_session)) -> List[schemas.UserAssignmentRow]:
    assignments = session.exec(
        select(models.UserAssignment)
        .options(
            selectinload(models.UserAssignment.user),
            selectinload(models.UserAssignment.administrative_unit),
        )
        .order_by(models.UserAssignment.created_at.desc())
    ).all()

    results: List[schemas.UserAssignmentRow] = []
    for assignment in assignments:
        unit = assignment.administrative_unit
        if unit is None:
            continue
        user = assignment.user
        if user is None:
            user_email = f"usuario{assignment.user_id}@desconocido.local"
            user_name = f"Usuario {assignment.user_id}"
            user_id = assignment.user_id
        else:
            user_email = user.email
            user_name = user.name
            user_id = user.id
        results.append(
            schemas.UserAssignmentRow(
                assignment_id=assignment.id,
                user_id=user_id,
                user_name=user_name,
                user_email=user_email,
                role=assignment.role,
                role_label=_role_label(assignment.role),
                administrative_unit_id=unit.id,
                administrative_unit_name=unit.name,
                administrative_unit_type=unit.unit_type,
                created_at=assignment.created_at,
            )
        )
    return results


@router.get("/committees", response_model=List[schemas.CommitteeDashboardOut])
def list_committees(session: Session = Depends(get_session)) -> List[schemas.CommitteeDashboardOut]:
    committees = _load_committees(session)
    return _serialize_committees(session, committees)


@router.get("/committees/{committee_id}", response_model=schemas.CommitteeDashboardOut)
def get_committee_detail(committee_id: int, session: Session = Depends(get_session)) -> schemas.CommitteeDashboardOut:
    committee = next(iter(_load_committees(session, committee_id)), None)
    if committee is None:
        raise HTTPException(status_code=404, detail="Comité no encontrado")
    serialized = _serialize_committees(session, [committee])
    if not serialized:
        raise HTTPException(status_code=404, detail="Comité no encontrado")
    return serialized[0]


@router.get("/exports/committees.xlsx")
def export_committees_excel(session: Session = Depends(get_session)) -> StreamingResponse:
    committees = _serialize_committees(session, _load_committees(session))
    workbook = Workbook()
    ws_committees = workbook.active
    ws_committees.title = "Comites"
    ws_committees.append(
        [
            "ID",
            "Nombre",
            "Tipo",
            "Sección",
            "Municipio",
            "Presidencia",
            "Teléfono",
            "Correo",
            "Total Integrantes",
            "Propietario",
            "Fecha de creación",
        ]
    )
    for committee in committees:
        municipality = committee.section.nombre_municipio if committee.section else ""
        ws_committees.append(
            [
                committee.id,
                committee.name,
                committee.type,
                committee.section_number,
                municipality,
                committee.presidente,
                committee.telefono,
                committee.email,
                committee.total_members,
                committee.owner_name or committee.owner_id,
                committee.created_at.strftime("%Y-%m-%d %H:%M"),
            ]
        )

    ws_members = workbook.create_sheet("Integrantes")
    ws_members.append([
        "ID Comité",
        "Comité",
        "Nombre integrante",
        "Teléfono",
        "Correo",
        "Sección",
        "Invitado por",
    ])
    for committee in committees:
        for member in committee.members:
            ws_members.append([
                committee.id,
                committee.name,
                member.full_name,
                member.phone,
                member.email,
                member.section_number,
                member.invited_by,
            ])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    filename = f"comites_r21_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/committees/{committee_id}/acta.pdf")
def download_committee_acta(committee_id: int, session: Session = Depends(get_session)) -> StreamingResponse:
    committee_model = next(iter(_load_committees(session, committee_id)), None)
    if committee_model is None:
        raise HTTPException(status_code=404, detail="Comité no encontrado")
    committee = _serialize_committees(session, [committee_model])[0]

    folio_seed = f"{committee.id}-{committee.created_at.isoformat()}-{committee.email}-{committee.total_members}"
    folio = hashlib.sha256(folio_seed.encode("utf-8")).hexdigest()[:12].upper()

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_title(_latin1(f"Acta Comité {committee.name}"))

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, _latin1("Acta de Comité R21 Michoacán"), ln=True, align="C")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 12)
    pdf.multi_cell(
        0,
        8,
        _latin1(
            "El Segundo Piso de la Cuarta Transformación R21 MORENA en Michoacán fortalece la organización territorial y la esperanza de nuestro movimiento."
        ),
    )
    pdf.ln(2)
    pdf.multi_cell(0, 8, _latin1(f"Comité: {committee.name}"))
    pdf.multi_cell(0, 8, _latin1(f"Presidencia: {committee.presidente} | Teléfono: {committee.telefono}"))

    if committee.section and committee.section.nombre_municipio:
        pdf.multi_cell(
            0,
            8,
            _latin1(
                f"Municipio: {committee.section.nombre_municipio} | Distrito: {committee.section.nombre_distrito or 'N/D'} | Sección: {committee.section_number}"
            ),
        )
    else:
        pdf.multi_cell(0, 8, _latin1(f"Sección: {committee.section_number}"))

    pdf.multi_cell(0, 8, _latin1(f"Correo del comité: {committee.email}"))
    pdf.multi_cell(0, 8, _latin1(f"Coordinación responsable: {committee.owner_name or committee.owner_id}"))
    pdf.multi_cell(0, 8, _latin1(f"Fecha de creación: {committee.created_at.strftime('%d/%m/%Y %H:%M')} UTC"))
    pdf.ln(4)

    pdf.set_font("Helvetica", "B", 12)
    pdf.cell(0, 8, _latin1("Integrantes registrados"), ln=True)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(80, 8, _latin1("Nombre"), border=1)
    pdf.cell(40, 8, _latin1("Teléfono"), border=1)
    pdf.cell(60, 8, _latin1("Correo"), border=1, ln=True)
    pdf.set_font("Helvetica", "", 10)

    if committee.members:
        for member in committee.members:
            pdf.cell(80, 8, _latin1(member.full_name)[:38], border=1)
            pdf.cell(40, 8, _latin1(member.phone)[:18], border=1)
            pdf.cell(60, 8, _latin1(member.email)[:30], border=1, ln=True)
    else:
        pdf.cell(180, 8, _latin1("Sin integrantes capturados"), border=1, ln=True, align="C")

    pdf.ln(6)
    pdf.set_font("Helvetica", "", 11)
    pdf.multi_cell(
        0,
        8,
        _latin1(
            "Este documento legitima el compromiso social del Segundo Piso de la Cuarta Transformación R21 MORENA en cada rincón de Michoacán."
        ),
    )
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 11)
    pdf.multi_cell(0, 8, _latin1(f"Folio electrónico: {folio}"))

    pdf_output = pdf.output(dest="S").encode("latin-1")
    buffer = BytesIO(pdf_output)
    buffer.seek(0)
    filename = f"acta_comite_{committee.id}.pdf"

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/documents", response_model=List[schemas.DocumentGalleryItem])
def list_documents(session: Session = Depends(get_session)) -> List[schemas.DocumentGalleryItem]:
    statement = (
        select(models.CommitteeDocument, models.Committee)
        .join(models.Committee, models.Committee.id == models.CommitteeDocument.committee_id)
        .order_by(models.CommitteeDocument.created_at.desc())
    )
    rows = session.exec(statement).all()
    items: List[schemas.DocumentGalleryItem] = []
    for document, committee in rows:
        items.append(
            schemas.DocumentGalleryItem(
                id=document.id,
                committee_id=committee.id,
                committee_name=committee.name,
                url=_normalize_upload_path(document.filename),
                original_name=document.original_name,
                content_type=document.content_type,
                size=document.size,
                created_at=document.created_at,
            )
        )
    return items


@router.get("/metrics", response_model=schemas.DashboardMetrics)
def dashboard_metrics(session: Session = Depends(get_session)) -> schemas.DashboardMetrics:
    total_committees = int(_safe_scalar(session, select(func.count(models.Committee.id)), 0))
    total_promovidos = int(_safe_scalar(session, select(func.count(models.CommitteeMember.id)), 0))
    total_documentos = int(_safe_scalar(session, select(func.count(models.CommitteeDocument.id)), 0))

    secciones_cubiertas = int(
        _safe_scalar(
            session,
            select(func.count(distinct(models.Committee.section_number))).where(
                models.Committee.section_number.is_not(None),
                models.Committee.section_number != "",
            ),
            0,
        )
    )
    total_secciones = int(_safe_scalar(session, select(func.count(models.Seccion.id)), 0))
    porcentaje_secciones = (
        round((secciones_cubiertas / total_secciones) * 100, 2) if total_secciones > 0 else 0.0
    )

    try:
        municipios_cubiertos = int(
            _safe_scalar(
                session,
                select(func.count(distinct(models.Seccion.nombre_municipio)))
                .select_from(models.Committee)
                .outerjoin(
                    models.Seccion,
                    cast(models.Seccion.id, String) == models.Committee.section_number,
                )
                .where(
                    models.Committee.section_number.is_not(None),
                    models.Committee.section_number != "",
                    models.Seccion.nombre_municipio.is_not(None),
                ),
                0,
            )
        )
    except SQLAlchemyError:
        municipios_cubiertos = 0

    porcentaje_municipios = (
        round(min(municipios_cubiertos / MUNICIPALITY_TARGET, 1.0) * 100, 2)
        if MUNICIPALITY_TARGET > 0
        else 0.0
    )

    return schemas.DashboardMetrics(
        total_committees=total_committees,
        total_promovidos=total_promovidos,
        municipios_cubiertos=municipios_cubiertos,
        municipios_meta=MUNICIPALITY_TARGET,
        porcentaje_municipios=porcentaje_municipios,
        secciones_cubiertas=secciones_cubiertas,
        total_secciones=total_secciones,
        porcentaje_secciones=porcentaje_secciones,
        total_documentos=total_documentos,
    )
